#!/usr/bin/env python3
# encoding: utf-8
'''
asap.prepareJSONInput -- Create a JSON input file for ASAP from a multifasta or Excel spreadsheet

asap.prepareJSONInput

@author:     Darrin Lemmer

@copyright:  2015 TGen North. All rights reserved.

@license:    ACADEMIC AND RESEARCH LICENSE -- see ../LICENSE

@contact:    dlemmer@tgen.org
'''

import sys
import os
import re
import argparse
import logging
import skbio.io
from skbio import DNA
from skbio import SequenceCollection
from openpyxl import load_workbook

import asap.dispatcher as dispatcher
import asap.assayInfo as assayInfo
from asap.assayInfo import SNP

__all__ = []
__version__ = 0.1
__date__ = '2015-08-03'
__updated__ = '2015-08-03'

DEBUG = 1
TESTRUN = 0
PROFILE = 0

class CLIError(Exception):
    '''Generic exception to raise and log different fatal errors.'''
    def __init__(self, msg):
        super(CLIError).__init__(type(self))
        self.msg = "E: %s" % msg
    def __str__(self):
        return self.msg
    def __unicode__(self):
        return self.msg

def main(argv=None): # IGNORE:C0111
    '''Command line options.'''

    if argv is None:
        argv = sys.argv
    else:
        sys.argv.extend(argv)

    program_name = os.path.basename(sys.argv[0])
    program_version = "v%s" % __version__
    program_build_date = str(__updated__)
    program_version_message = '%%(prog)s %s (%s)' % (program_version, program_build_date)
    if __name__ == '__main__':
        program_shortdesc = __import__('__main__').__doc__.split("\n")[1]
    else:
        program_shortdesc = __doc__.split("\n")[1]    
    program_license = '''%s

  Created by TGen North on %s.
  Copyright 2015 TGen North. All rights reserved.

  Available for academic and research use only under a license
  from The Translational Genomics Research Institute (TGen)
  that is free for non-commercial use.

  Distributed on an "AS IS" basis without warranties
  or conditions of any kind, either express or implied.

USAGE
''' % (program_shortdesc, str(__date__))

    try:
        # Setup argument parser
        parser = argparse.ArgumentParser(description=program_license, formatter_class=argparse.RawDescriptionHelpFormatter)
        required_group = parser.add_argument_group("required arguments")
        exclusive_group = required_group.add_mutually_exclusive_group(required=True)
        exclusive_group.add_argument("-f", "--fasta", metavar="FILE", help="fasta file containing amplicon sequences.")
        exclusive_group.add_argument("-x", "--excel", metavar="FILE", help="Excel file of assay data.")
        required_group.add_argument("-o", "--out", metavar="FILE", required=True, help="output JSON file to write. [REQUIRED]")
        parser.add_argument('-V', '--version', action='version', version=program_version_message)

        # Process arguments
        args = parser.parse_args()

        fasta_file = args.fasta
        excel_file = args.excel
        out_file = args.out

        assay_list = []

        if fasta_file:
            sc = skbio.io.registry.read(fasta_file, format='fasta', into=SequenceCollection, constructor=DNA)
            for seq in sc:
                significance = assayInfo.Significance(seq.metadata['description']) if seq.metadata['description'] else None
                amplicon = assayInfo.Amplicon(sequence=str(seq), significance=significance)
                target = assayInfo.Target(function='species ID', amplicon=amplicon)
                assay = assayInfo.Assay(name=seq.metadata['id'], assay_type='presence/absence', target=target)
                assay_list.append(assay)
        else:
            wb = load_workbook(excel_file, read_only=True)
            ws = wb.active
            amplicon = None
            target = None
            assay = None
            for row in ws.iter_rows(row_offset=2):
                
                if row[0].value: #Create a new assay
                    if assay:
                        assay_list.append(assay)
                        target = None
                        amplicon = None
                    assay = assayInfo.Assay(name=row[0].value, assay_type=row[1].value)
                
                significance = assayInfo.Significance(message=row[15].value)
                element = None
                if row[12].value: #Significance gets attached to ROI
                    element = assayInfo.RegionOfInterest(position_range=row[12].value, aa_sequence=row[13].value, significance=significance)
                elif row[9].value: #Significance gets attached to SNP
                    element = assayInfo.SNP(position=row[9].value, reference=row[10].value, variant=row[11].value, significance=significance)
                if row[8].value: #Process amplicon
                    amplicon = assayInfo.Amplicon(sequence=row[8].value, variant_name=row[7].value)
                if element:
                    amplicon.add_SNP(element) if isinstance(element, assayInfo.SNP) else amplicon.add_ROI(element)
                else:
                    amplicon.significance = significance
                
                if target:
                    target.add_amplicon(amplicon)
                    amplicon = None
                else:
                    target = assayInfo.Target(function=row[2].value, gene_name=row[3].value, start_position=row[4].value, end_position=row[5].value, reverse_comp=row[6].value, amplicon=amplicon)
                    assay.target = target
                                                
        assay_list.append(assay) #get the last one
        assay_data = {"Assay":assay_list} 
        assayInfo.writeJSON(assay_data, out_file)

        return 0
    except KeyboardInterrupt:
        ### handle keyboard interrupt ###
        return 0
    except Exception as e:
        if DEBUG or TESTRUN:
            raise(e)
        indent = len(program_name) * " "
        sys.stderr.write(program_name + ": " + repr(e) + "\n")
        sys.stderr.write(indent + "  for help use --help")
        return 2

if __name__ == "__main__":
    if DEBUG:
        pass
    if TESTRUN:
        import doctest
        doctest.testmod()
    if PROFILE:
        import cProfile
        import pstats
        profile_filename = 'asap.outputCombiner_profile.txt'
        cProfile.run('main()', profile_filename)
        statsfile = open("profile_stats.txt", "wb")
        p = pstats.Stats(profile_filename, stream=statsfile)
        stats = p.strip_dirs().sort_stats('cumulative')
        stats.print_stats()
        statsfile.close()
        sys.exit(0)
    sys.exit(main())